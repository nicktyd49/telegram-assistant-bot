from __future__ import annotations

import copy
import logging
import math
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.colors import Color

from config import settings
from services import onedrive_service

logger = logging.getLogger("assistant-bot.policy_workbook")

CLIENT_DIR = Path(__file__).resolve().parent.parent / "client_policy_summaries"
SHEET_NAME = "Policy Summary"

# Where client workbooks live on OneDrive once ONEDRIVE_CLIENT_ID is set up —
# this is the actual persistent copy; CLIENT_DIR above is just a working
# cache on whatever disk the bot happens to be running on (which, on
# Railway, gets wiped on every redeploy).
ONEDRIVE_WORKBOOK_FOLDER = "Client"

HEADER_ROW = 4
FIRST_DATA_ROW = 5
TOTAL_ROW = 12
LAST_DATA_COLUMN = 17  # column Q — the table stops at Remarks, matching the reference sheet

DATE_FMT = "dd/mm/yyyy"
CURRENCY_FMT_2DP = '"$"#,##0.00'
CURRENCY_FMT_0DP = '"$"#,##0'

# Styling matched to the agent's real reference sheet (Policy Summary Johnson):
# light-blue header band, Times New Roman throughout, thin borders on every
# cell, and light alternating row banding.
HEADER_FILL = PatternFill(start_color="FF99CCFF", end_color="FF99CCFF", fill_type="solid")
BAND_FILL = PatternFill(fgColor=Color(theme=0, tint=-0.1499984740745262), fill_type="solid")
TOTAL_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
NO_FILL = PatternFill(fill_type=None)
THIN = Side(style="thin")
CELL_BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

TITLE_FONT = Font(name="Times New Roman", size=20)
LABEL_FONT = Font(name="Times New Roman", size=14)
CLIENT_NAME_FONT = Font(name="Times New Roman", size=14, bold=True, color="FFFF0000")
DOB_FONT = Font(name="Times New Roman", size=12, bold=True)
HEADER_FONT = Font(name="Times New Roman", size=11, bold=True)
DATA_FONT = Font(name="Times New Roman", size=11)
TOTAL_FONT = Font(name="Times New Roman", size=12, bold=True)
FOOTER_FONT = Font(name="Times New Roman", size=10)

CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

COLUMN_HEADERS = {
    "A": "No",
    "B": "Company",
    "C": "Policy No",
    "D": "Payment Date",
    "E": "Plan Type",
    "F": "Premium Annual\n(Cash)",
    "G": "Premium Annual\n(CPF)",
    "H": "Payment Frequency",
    "I": "Mode of Payment",
    "J": "Total Death Coverage",
    "K": "Total Permanent Disability Coverage",
    "L": "Critical Illness Coverage",
    "M": "Early Stage Illness Coverage",
    "N": "Disability Income\n(Per Mth)",
    "O": "Total Accident - Lump Sum",
    "P": "Total Accident - Medical Reimbursement",
    "Q": "Remarks",
}

COLUMN_WIDTHS = {
    "A": 4.17, "B": 10.63, "C": 16, "D": 13, "E": 13, "F": 13, "G": 12.91,
    "H": 11.03, "I": 10.49, "J": 15.33, "K": 14.39, "L": 10.36, "M": 11.43,
    "N": 9.14, "O": 12.51, "P": 15.87, "Q": 26,
}

CURRENCY_2DP_COLUMNS = ["F", "G"]
CURRENCY_0DP_COLUMNS = ["J", "K", "L", "M", "N", "O", "P"]
NUMERIC_COLUMNS = CURRENCY_2DP_COLUMNS + CURRENCY_0DP_COLUMNS

# C (Policy No) and Q (Remarks) are built/handled separately below — C combines
# plan_name + policy_no, and coverage_end_age no longer gets its own column
# (it's still used in-memory to work out the Payment Date range, just not
# persisted as a visible column — matches the agent's reference sheet exactly).
COLUMN_FIELDS = {
    "B": "company",
    "E": "plan_type",
    "F": "premium_annual_cash",
    "G": "premium_annual_cpf",
    "H": "payment_frequency",
    "I": "mode_of_payment",
    "J": "total_death_coverage",
    "K": "total_permanent_disability_coverage",
    "L": "critical_illness_coverage",
    "M": "early_stage_illness_coverage",
    "N": "disability_income_per_month",
    "O": "total_accident_lump_sum",
    "P": "total_accident_medical_reimbursement",
    "Q": "remarks",
}

# Hidden helper columns past the visible table (R onward) - not part of the
# agent-facing Policy Summary layout, not styled/bordered, not summed into
# the Total row. They exist purely so the Policy Illustration sheet has real
# data to plot instead of guessing: coverage end age (for the age axis and a
# policy's bar length), an investment plan's surrender values, and - if a
# policy's cover is documented as stepping down at a given age - the age and
# amount it drops to, so the illustration bar can shade that portion
# differently instead of implying full coverage runs the whole way through.
HIDDEN_COLUMN_FIELDS = {
    "R": "surrender_value_current",
    "S": "surrender_value_at_65",
    "T": "coverage_end_age",
    "U": "coverage_drop_age",
    "V": "reduced_coverage_amount",
}


class PolicyWorkbookError(RuntimeError):
    pass


_ONEDRIVE_ILLEGAL_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')

WORKBOOK_FILENAME_PREFIX = "Policy summary "


def _onedrive_safe_name(text: str | None, fallback: str | None = None) -> str | None:
    """Strips only the characters OneDrive/Windows actually forbid in a file
    or folder name, keeping spaces and casing intact — so this matches
    whatever Nic already has saved by hand on OneDrive (e.g. "Tan Zhen
    Xuan", "Policy summary Johnson.xlsx") instead of underscore-mangling
    names that were never actually a problem."""
    cleaned = _ONEDRIVE_ILLEGAL_CHARS.sub("", (text or "")).strip().rstrip(".")
    return cleaned or fallback


def _safe_filename(client_name: str) -> str:
    return f"{WORKBOOK_FILENAME_PREFIX}{_onedrive_safe_name(client_name, 'Unknown Client')}.xlsx"


def _client_path(client_name: str) -> Path:
    CLIENT_DIR.mkdir(parents=True, exist_ok=True)
    path = CLIENT_DIR / _safe_filename(client_name)
    _sync_from_onedrive(client_name, path)
    return path


def _onedrive_remote_path(client_name: str, path: Path) -> str:
    folder = _onedrive_safe_name(client_name, "Unknown_Client")
    return f"{ONEDRIVE_WORKBOOK_FOLDER}/{folder}/{path.name}"


def _sync_from_onedrive(client_name: str, path: Path) -> None:
    """Pulls the latest saved copy of this client's workbook down from
    OneDrive before we touch it locally, so we're always editing on top of
    the real persisted version rather than whatever (possibly stale, or
    wiped-by-redeploy) copy happens to be sitting on local disk. Silently
    does nothing if OneDrive isn't configured yet, or if this client simply
    doesn't have a workbook on OneDrive yet (a brand-new client)."""
    if not settings.onedrive_configured:
        return
    try:
        data = onedrive_service._download_bytes_sync(_onedrive_remote_path(client_name, path))
    except Exception:  # noqa: BLE001
        logger.exception(
            "Could not check OneDrive for %s — continuing with whatever local copy is on disk",
            path.name,
        )
        return
    if data is not None:
        path.write_bytes(data)


def _sync_to_onedrive(client_name: str, path: Path) -> None:
    """Pushes the just-saved local workbook up to OneDrive so it survives a
    redeploy. Non-fatal if it fails — the caller already has a good local
    copy to send back to Nic; it just won't be backed up until the next
    successful save."""
    if not settings.onedrive_configured:
        return
    try:
        onedrive_service._upload_bytes_sync(_onedrive_remote_path(client_name, path), path.read_bytes())
    except Exception:  # noqa: BLE001
        logger.exception(
            "Failed to sync %s to OneDrive — local copy is saved, but it won't survive a redeploy "
            "until this succeeds",
            path.name,
        )


async def list_client_names() -> list[str]:
    # All client folder names currently on OneDrive under Client/ - read from
    # OneDrive directly (not the local cache in CLIENT_DIR) since that is the
    # actual persisted list; local disk gets wiped on every Railway redeploy
    # so it is not reliable as an index of "who are my clients".
    if not settings.onedrive_configured:
        return []
    children = await onedrive_service.list_children(ONEDRIVE_WORKBOOK_FOLDER)
    return sorted(c["name"] for c in children if "folder" in c)


def _load_client_workbook(client_name: str):
    # Read-only load of a client's existing workbook - unlike _open_or_create,
    # this does NOT build a blank workbook when one does not exist yet; a
    # lookup command has nothing useful to show for a client with no policies
    # logged, so it should say so rather than showing an empty sheet.
    path = _client_path(client_name)
    if not path.exists():
        raise PolicyWorkbookError(f"No policy summary found yet for {client_name}.")
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise PolicyWorkbookError(f"Workbook is missing a '{SHEET_NAME}' sheet")
    return wb


def get_client_summary(client_name: str) -> dict:
    # Everything worth showing about a client in a quick chat lookup: their
    # policies (one dict per row), the totals row, and the same coverage-gap
    # notes add_policy_row already computes on every save.
    wb = _load_client_workbook(client_name)
    ws = wb[SHEET_NAME]
    total_row = _find_total_row(ws)
    last_row = total_row - 1

    dob = ws["K3"].value
    dob_display = _fmt_date_display(dob) if isinstance(dob, date) else (dob or None)
    age_next_birthday = None
    if isinstance(dob, date):
        # Same arithmetic as the live Q3 formula: DATEDIF(K3,TODAY(),"y")+1.
        today = date.today()
        completed_years = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
        age_next_birthday = completed_years + 1

    policies = []
    for row in range(FIRST_DATA_ROW, total_row):
        company = ws.cell(row=row, column=2).value  # B
        if company in (None, ""):
            continue
        policies.append({
            "company": company,
            "policy_no": ws.cell(row=row, column=3).value,  # C
            "payment_date": ws.cell(row=row, column=4).value,  # D
            "plan_type": ws.cell(row=row, column=5).value,  # E
            "premium_cash": ws.cell(row=row, column=6).value,  # F
            "premium_cpf": ws.cell(row=row, column=7).value,  # G
            "payment_frequency": ws.cell(row=row, column=8).value,  # H
            "mode_of_payment": ws.cell(row=row, column=9).value,  # I
            "death_coverage": ws.cell(row=row, column=10).value,  # J
            "tpd_coverage": ws.cell(row=row, column=11).value,  # K
            "ci_coverage": ws.cell(row=row, column=12).value,  # L
            "early_stage_coverage": ws.cell(row=row, column=13).value,  # M
            "di_coverage": ws.cell(row=row, column=14).value,  # N
            "accident_lump_sum": ws.cell(row=row, column=15).value,  # O
            "accident_medical": ws.cell(row=row, column=16).value,  # P
            "remarks": ws.cell(row=row, column=17).value,  # Q
        })

    totals = {
        "premium_cash": ws[f"F{total_row}"].value,
        "premium_cpf": ws[f"G{total_row}"].value,
        "death_coverage": ws[f"J{total_row}"].value,
        "tpd_coverage": ws[f"K{total_row}"].value,
        "ci_coverage": ws[f"L{total_row}"].value,
        "early_stage_coverage": ws[f"M{total_row}"].value,
        "di_coverage": ws[f"N{total_row}"].value,
        "accident_lump_sum": ws[f"O{total_row}"].value,
        "accident_medical": ws[f"P{total_row}"].value,
    }

    action_items = action_items_for(ws, total_row, FIRST_DATA_ROW, last_row)

    return {
        "client_name": client_name,
        "date_of_birth": dob_display,
        "age_next_birthday": age_next_birthday,
        "policies": policies,
        "totals": totals,
        "action_items": action_items,
    }


def _to_number_or_text(value):
    """Most coverage/premium fields are plain numbers. Some investment-linked
    plans express a benefit as a formula instead of a fixed sum (e.g. "101% of
    total premiums paid") — keep that as descriptive text rather than mangling
    it into a meaningless partial number."""
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return value
    s = str(value).strip()
    cleaned = re.sub(r"[^0-9.\-]", "", s)
    if cleaned and cleaned not in ("-", ".") and cleaned == s.replace(",", ""):
        try:
            return float(cleaned)
        except ValueError:
            pass
    return s


def _to_date_or_text(value):
    """Parses common date strings into a real date object (so Excel treats it
    as a date and can be formatted/sorted). Falls back to the original text
    if it doesn't parse cleanly."""
    if value in (None, ""):
        return None
    if isinstance(value, (datetime, date)):
        return value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return s


def _ensure_age_formula(ws) -> None:
    """Age Next Birthday (Q3) is a live Excel formula off Date of Birth (K3),
    not a value computed once by the bot — so if DOB isn't in any uploaded
    document (common — many policy schedules only state an age, not a literal
    birth date) the agent can just type it into K3 directly in Excel and Q3
    updates itself immediately, no re-run needed."""
    ws["Q3"] = '=IF(K3="","",DATEDIF(K3,TODAY(),"y")+1)'
    ws["Q3"].font = DOB_FONT


def _parse_date_loose(value):
    """Best-effort parse of a date-ish value into a date object, or None."""
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _fmt_date_display(d) -> str:
    return d.strftime("%d-%b-%Y")


def _add_years(d: date, years: int) -> date:
    try:
        return d.replace(year=d.year + years)
    except ValueError:
        # e.g. dob is Feb 29 and the target year isn't a leap year
        return d.replace(year=d.year + years, day=28)


def _compute_maturity_date(maturity_raw, coverage_end_age, dob_date: date | None, plan_type):
    explicit = _parse_date_loose(maturity_raw)
    if explicit is not None:
        return explicit
    if dob_date is not None and isinstance(coverage_end_age, (int, float)):
        return _add_years(dob_date, int(coverage_end_age))
    if dob_date is not None and isinstance(plan_type, str) and plan_type.strip().lower() == "wol":
        # Whole-life plans have no stated end date — model them out to age 99,
        # the standard actuarial convention for "for life" illustrations.
        return _add_years(dob_date, 99)
    return None


def _build_date_range_text(fields: dict, dob_date: date | None) -> str | None:
    start = _parse_date_loose(fields.get("policy_start_date"))
    maturity = _compute_maturity_date(
        fields.get("maturity_date"), fields.get("coverage_end_age"), dob_date, fields.get("plan_type")
    )
    if start is None and maturity is None:
        return None
    start_text = _fmt_date_display(start) if start else "?"
    maturity_text = _fmt_date_display(maturity) if maturity else "?"
    return f"{start_text} - {maturity_text}"


def _build_policy_no_text(fields: dict) -> str | None:
    """Policy No column shows the product's actual plan name together with
    its policy number, e.g. "LifeReady 25\\n2491554220", so the sheet reads
    clearly without needing a separate column."""
    plan_name = fields.get("plan_name")
    policy_no = fields.get("policy_no")
    plan_name = str(plan_name).strip() if plan_name not in (None, "") else None
    policy_no = str(policy_no).strip() if policy_no not in (None, "") else None
    if plan_name and policy_no:
        return f"{plan_name}\n{policy_no}"
    return plan_name or policy_no


def _backfill_maturity_dates(ws, dob_date: date | None, skip_row: int, total_row: int) -> None:
    """If a client's DOB only becomes known partway through adding their
    policies, earlier rows that were left showing "... - ?" for maturity
    (because we didn't know their age yet) can now be worked out for the
    cases that don't need a persisted end-age (an explicit date, or a WOL
    plan modelled out to age 99). Re-checks every existing row."""
    if dob_date is None:
        return
    for r in range(FIRST_DATA_ROW, total_row):
        if r == skip_row:
            continue
        current = ws.cell(row=r, column=4).value  # column D
        if not isinstance(current, str) or not current.endswith("- ?"):
            continue
        start_text = current.rsplit(" - ", 1)[0]
        plan_type = ws.cell(row=r, column=5).value  # column E
        maturity = _compute_maturity_date(None, None, dob_date, plan_type)
        if maturity is not None:
            ws.cell(row=r, column=4).value = f"{start_text} - {_fmt_date_display(maturity)}"


def _estimate_wrapped_lines(text, col_width: float) -> int:
    """Rough estimate of how many display lines a wrapped cell will need,
    given the column's character width, so row height can expand to fit
    longer Remarks / Policy No content instead of clipping it."""
    if text in (None, ""):
        return 1
    chars_per_line = max(int(col_width * 1.1), 6)
    lines = 0
    for segment in str(text).split("\n"):
        lines += max(1, math.ceil(len(segment) / chars_per_line))
    return lines


def _autofit_row_height(ws, row: int) -> None:
    """Expands a data row's height when its Remarks or Policy No content
    needs more room to display neatly, matching how the reference sheet's
    rows grow for longer notes instead of staying visually cramped."""
    q_lines = _estimate_wrapped_lines(ws.cell(row=row, column=17).value, COLUMN_WIDTHS["Q"])
    c_lines = _estimate_wrapped_lines(ws.cell(row=row, column=3).value, COLUMN_WIDTHS["C"])
    lines = max(q_lines, c_lines, 3)
    ws.row_dimensions[row].height = max(56.25, lines * 18)


def _style_data_row(ws, row: int) -> None:
    """Applies the Johnson-matched look to one data row: Times New Roman,
    thin border box, centered+wrapped, and light alternating row banding."""
    banded = (row - FIRST_DATA_ROW) % 2 == 1
    fill = BAND_FILL if banded else NO_FILL
    for col in range(1, LAST_DATA_COLUMN + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = DATA_FONT
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True) if col != 17 \
            else Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.fill = fill
    for col in CURRENCY_2DP_COLUMNS:
        ws[f"{col}{row}"].number_format = CURRENCY_FMT_2DP
    for col in CURRENCY_0DP_COLUMNS:
        ws[f"{col}{row}"].number_format = CURRENCY_FMT_0DP
    _autofit_row_height(ws, row)


def _style_total_row(ws, total_row: int) -> None:
    """Yellow-fills the Total row across the table width so it's visually
    unmistakable from the data rows above it."""
    for col in range(1, LAST_DATA_COLUMN + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.font = TOTAL_FONT
        cell.border = CELL_BORDER
        cell.fill = TOTAL_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _build_new_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    ws["I1"] = "Policy Summary"
    ws["I1"].font = TITLE_FONT

    ws["A3"] = "Scope of Aggregated Coverage for"
    ws["A3"].font = LABEL_FONT
    ws["J3"] = "Date of Birth:"
    ws["J3"].font = LABEL_FONT
    ws["P3"] = "Age Next Birthday: "
    ws["P3"].font = LABEL_FONT
    _ensure_age_formula(ws)

    for col, header in COLUMN_HEADERS.items():
        cell = ws[f"{col}{HEADER_ROW}"]
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = CELL_BORDER
        cell.alignment = CENTER_WRAP

    for row in range(FIRST_DATA_ROW, TOTAL_ROW):
        _style_data_row(ws, row)

    ws[f"C{TOTAL_ROW}"] = "Total"
    _style_total_row(ws, TOTAL_ROW)
    for col in CURRENCY_2DP_COLUMNS:
        ws[f"{col}{TOTAL_ROW}"].number_format = CURRENCY_FMT_2DP
    for col in CURRENCY_0DP_COLUMNS:
        ws[f"{col}{TOTAL_ROW}"].number_format = CURRENCY_FMT_0DP

    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[HEADER_ROW].height = 60
    ws.freeze_panes = "A5"

    return wb


def _open_or_create(client_name: str):
    path = _client_path(client_name)
    if path.exists():
        wb = openpyxl.load_workbook(path)
    else:
        wb = _build_new_workbook()
    if SHEET_NAME not in wb.sheetnames:
        raise PolicyWorkbookError(f"Workbook is missing a '{SHEET_NAME}' sheet")
    return wb, path


def _find_total_row(ws) -> int:
    for row in range(1, ws.max_row + 2):
        if str(ws.cell(row=row, column=3).value or "").strip().lower() == "total":
            return row
    return TOTAL_ROW


def _find_data_row(ws, total_row: int):
    for row in range(FIRST_DATA_ROW, total_row):
        if ws.cell(row=row, column=2).value in (None, ""):
            return row, False
    return total_row, True


def _copy_row_style(ws, src_row: int, dst_row: int) -> None:
    for col in range(1, LAST_DATA_COLUMN + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        dst.number_format = src.number_format
        dst.font = copy.copy(src.font)
        dst.border = copy.copy(src.border)
        dst.fill = copy.copy(src.fill)
        dst.alignment = copy.copy(src.alignment)


def _recompute_total(ws, total_row: int, first_row: int, last_row: int) -> None:
    for col in NUMERIC_COLUMNS:
        total = 0.0
        for row in range(first_row, last_row + 1):
            val = ws[f"{col}{row}"].value
            if isinstance(val, (int, float)):
                total += val
        ws[f"{col}{total_row}"] = total


def _clear_old_footer(ws, search_rows: int) -> None:
    for row in range(1, search_rows + 1):
        a_cell = ws.cell(row=row, column=1)
        if isinstance(a_cell.value, str) and a_cell.value.startswith("Updated as of"):
            a_cell.value = None
        p_cell = ws.cell(row=row, column=16)
        if isinstance(p_cell.value, str) and p_cell.value.startswith("Prepared by"):
            p_cell.value = None


def _write_footer(ws, total_row: int) -> None:
    _clear_old_footer(ws, total_row + 6)
    footer_row = total_row + 2
    a_cell = ws.cell(row=footer_row, column=1, value=f"Updated as of {date.today().strftime('%B %Y')}")
    a_cell.font = FOOTER_FONT
    p_cell = ws.cell(row=footer_row, column=16, value=f"Prepared by {settings.agent_name}")
    p_cell.font = FOOTER_FONT


def add_policy_row(client_name: str, fields: dict, date_of_birth: str | None = None):
    """Adds one policy row for client_name (creating their workbook if needed),
    recomputes the Total row, and saves. Returns (path, policy_count)."""
    wb, path = _open_or_create(client_name)
    ws = wb[SHEET_NAME]
    # Johnson's reference sheet has gridlines turned off (a clean white canvas
    # instead of Excel's default checkerboard) - matched here so every client
    # workbook looks the same, including ones created before this was added.
    ws.sheet_view.showGridLines = False

    if not ws["F3"].value:
        ws["F3"] = client_name
        ws["F3"].font = CLIENT_NAME_FONT
    dob = date_of_birth or fields.get("date_of_birth") or ws["K3"].value
    if (date_of_birth or fields.get("date_of_birth")) and not ws["K3"].value:
        dob_value = _to_date_or_text(date_of_birth or fields.get("date_of_birth"))
        ws["K3"] = dob_value
        ws["K3"].font = DOB_FONT
        if isinstance(dob_value, date):
            ws["K3"].number_format = DATE_FMT
    _ensure_age_formula(ws)

    dob_date = _parse_date_loose(dob)

    total_row = _find_total_row(ws)
    row, needs_insert = _find_data_row(ws, total_row)

    if needs_insert:
        ws.insert_rows(row)
        _copy_row_style(ws, row - 1, row)
        total_row += 1

    for col, field_key in COLUMN_FIELDS.items():
        raw_value = fields.get(field_key)
        if col in NUMERIC_COLUMNS:
            ws[f"{col}{row}"] = _to_number_or_text(raw_value)
        else:
            ws[f"{col}{row}"] = raw_value

    for col, field_key in HIDDEN_COLUMN_FIELDS.items():
        raw_value = fields.get(field_key)
        if raw_value not in (None, ""):
            ws[f"{col}{row}"] = raw_value

    ws[f"C{row}"] = _build_policy_no_text(fields)
    ws[f"D{row}"] = _build_date_range_text(fields, dob_date)
    _backfill_maturity_dates(ws, dob_date, skip_row=row, total_row=total_row)

    ws[f"A{row}"] = row - FIRST_DATA_ROW + 1
    _style_data_row(ws, row)

    last_row = total_row - 1
    _recompute_total(ws, total_row, FIRST_DATA_ROW, last_row)
    _write_footer(ws, total_row)
    gap_notes = _coverage_gap_notes(ws, total_row, FIRST_DATA_ROW, last_row)

    wb.save(path)
    _sync_to_onedrive(client_name, path)
    policy_count = sum(
        1 for r in range(FIRST_DATA_ROW, total_row)
        if ws.cell(row=r, column=2).value not in (None, "")
    )
    return path, policy_count, gap_notes



# Rough rule-of-thumb threshold for flagging Critical Illness coverage as
# "thin" relative to Death coverage - not a rigid standard, just something
# worth a look/conversation with the client.
CI_THIN_RATIO = 0.2


def _column_has_coverage(ws, col: str, first_row: int, last_row: int) -> bool:
    """True if any data row in this column has a real value - a positive
    number, or a non-empty string (a formula-based benefit like "101% of
    premiums paid", which still counts as coverage even though it doesn't
    sum into the Total row)."""
    for row in range(first_row, last_row + 1):
        val = ws[f"{col}{row}"].value
        if isinstance(val, (int, float)) and val > 0:
            return True
        if isinstance(val, str) and val.strip():
            return True
    return False


def _coverage_flags(ws, total_row: int, first_row: int, last_row: int) -> dict:
    # Shared building block for _coverage_gap_notes (flat observations) and
    # action_items_for (the same gaps phrased as recommended next actions
    # for the Action Plan sheet / Telegram message) - keeps both in sync
    # instead of duplicating the coverage-column checks twice.
    if last_row < first_row:
        return {
            "has_death": False, "has_tpd": False, "has_ci": False,
            "has_early": False, "has_di": False, "death_total": 0, "ci_total": 0,
        }

    has_death = _column_has_coverage(ws, "J", first_row, last_row)
    has_tpd = _column_has_coverage(ws, "K", first_row, last_row)
    has_ci = _column_has_coverage(ws, "L", first_row, last_row)
    has_early = _column_has_coverage(ws, "M", first_row, last_row)
    has_di = _column_has_coverage(ws, "N", first_row, last_row)

    death_total = ws[f"J{total_row}"].value
    ci_total = ws[f"L{total_row}"].value
    death_total = death_total if isinstance(death_total, (int, float)) else 0
    ci_total = ci_total if isinstance(ci_total, (int, float)) else 0

    return {
        "has_death": has_death, "has_tpd": has_tpd, "has_ci": has_ci,
        "has_early": has_early, "has_di": has_di,
        "death_total": death_total, "ci_total": ci_total,
    }


def _coverage_gap_notes(ws, total_row: int, first_row: int, last_row: int) -> list[str]:
    """Short, data-driven observations about obvious coverage gaps across ALL
    of this client's policies on file - meant as talking points for the
    agent's next conversation with the client, not a definitive assessment."""
    f = _coverage_flags(ws, total_row, first_row, last_row)

    notes = []
    if f["has_death"] and not f["has_ci"] and not f["has_early"]:
        notes.append("No Critical Illness coverage on file.")
    if f["has_death"] and not f["has_tpd"]:
        notes.append("No TPD coverage on file.")
    if f["has_death"] and not f["has_di"]:
        notes.append("No Disability Income coverage on file.")
    if f["has_ci"] and f["death_total"] > 0 and f["ci_total"] > 0 and f["ci_total"] < CI_THIN_RATIO * f["death_total"]:
        notes.append(
            f"CI coverage (${f['ci_total']:,.0f}) looks thin next to Death coverage (${f['death_total']:,.0f})."
        )

    return notes[:3]


def action_items_for(ws, total_row: int, first_row: int, last_row: int) -> list[dict]:
    """Same coverage-gap detection as _coverage_gap_notes, phrased instead as
    {"gap": ..., "action": ...} - what's actually shown on the Action Plan
    sheet and sent to Nic on Telegram after a policy is logged. Always
    returns at least one item, so there is never a case with "gaps found"
    but no stated next step."""
    f = _coverage_flags(ws, total_row, first_row, last_row)

    items = []
    if f["has_death"] and not f["has_ci"] and not f["has_early"]:
        items.append({
            "gap": "No Critical Illness coverage on file.",
            "action": "Discuss adding Critical Illness coverage at the next review.",
        })
    if f["has_death"] and not f["has_tpd"]:
        items.append({
            "gap": "No TPD coverage on file.",
            "action": "Discuss adding Total & Permanent Disability coverage.",
        })
    if f["has_death"] and not f["has_di"]:
        items.append({
            "gap": "No Disability Income coverage on file.",
            "action": "Discuss adding Disability Income coverage to protect their monthly income.",
        })
    if f["has_ci"] and f["death_total"] > 0 and f["ci_total"] > 0 and f["ci_total"] < CI_THIN_RATIO * f["death_total"]:
        items.append({
            "gap": f"CI coverage (${f['ci_total']:,.0f}) looks thin next to Death coverage (${f['death_total']:,.0f}).",
            "action": "Review whether Critical Illness coverage should be increased.",
        })
    if not items:
        items.append({
            "gap": "No obvious coverage gaps found.",
            "action": "No urgent action - consider a routine check-in for life changes (marriage, kids, income change).",
        })

    return items[:4]

