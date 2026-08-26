from __future__ import annotations

from datetime import date, datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from services.policy_workbook import PolicyWorkbookError, SHEET_NAME, _client_path

# This sheet is a big-picture protection-vs-savings snapshot, not a home for
# every policy type. Only core life protection (Term/Whole Life) goes in the
# top "SUM ASSURED" half, and only investment-linked plans go in the bottom
# "CASH VALUE" half. Riders like PA/Health/CI/DI standalone plans still live
# in Policy Summary but are deliberately left off this chart.
PROTECTION_PLAN_TYPES = {"wol", "term"}
INVESTMENT_PLAN_TYPES = {"invest"}

BAND_FILL_PROTECTION = PatternFill(start_color="FFB1C2D9", end_color="FFB1C2D9", fill_type="solid")
BAND_FILL_PROTECTION_ALT = PatternFill(start_color="FFC9D3C4", end_color="FFC9D3C4", fill_type="solid")
BAND_FILL_PROTECTION_REDUCED = PatternFill(start_color="FFE0E7F0", end_color="FFE0E7F0", fill_type="solid")
BAND_FILL_PROTECTION_ALT_REDUCED = PatternFill(start_color="FFE6EBE3", end_color="FFE6EBE3", fill_type="solid")
BAND_FILL_INVESTMENT = PatternFill(start_color="FFF2EFD6", end_color="FFF2EFD6", fill_type="solid")
NO_FILL = PatternFill(fill_type=None)
SPINE_BORDER = Border(left=Side(style="thin", color="FF808080"))

# The Johnson reference sheet shades its CPF payout description in a dark
# navy font (rather than the sheet's default black) to make that specific
# line read as important at a glance. Matched exactly here.
IMPORTANT_TEXT_FONT = Font(name="Calibri", size=11, color="FF16365C")

BLOCK_WIDTH_COLS = 10  # columns A..J get the block's shading band

# CPF LIFE Standard Plan estimated monthly payout at 65, for a member
# meeting the Full Retirement Sum ($220,400) - matches this sheet's own
# "assuming meeting Minimum Sum Scheme" wording (FRS is the modern name for
# what used to be called the Minimum Sum). Source: CPF Board's own published
# figure for members turning 55 in 2026 (cpf.gov.sg), male member, Standard
# Plan. This is a today's-dollars planning estimate, not a personalized
# projection - CPF only publishes payout figures for members nearing 55/65,
# so a client decades away from retirement has no real figure to quote yet.
# CPF revises these numbers most years - re-check cpf.gov.sg and update this
# constant when that happens, rather than letting it go stale.
CPF_LIFE_FRS_MONTHLY_ESTIMATE = 1780


def _lane_for(plan_type) -> str:
    key = plan_type.strip().lower() if isinstance(plan_type, str) else ""
    if key in INVESTMENT_PLAN_TYPES:
        return "investment"
    if key in PROTECTION_PLAN_TYPES:
        return "protection"
    return "other"


def _current_age(dob) -> int | None:
    if dob in (None, ""):
        return None
    if isinstance(dob, datetime):
        dob_date = dob.date()
    elif isinstance(dob, date):
        dob_date = dob
    else:
        dob_date = None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                dob_date = datetime.strptime(str(dob), fmt).date()
                break
            except ValueError:
                continue
    if dob_date is None:
        return None
    today = date.today()
    age = today.year - dob_date.year
    if (today.month, today.day) < (dob_date.month, dob_date.day):
        age -= 1
    return age


def _illustration_sheet_name(client_name: str) -> str:
    prefix = "Policy Illustration "
    base = f"{prefix}{client_name}"
    if len(base) <= 31:
        return base
    return (prefix + client_name)[:31].rstrip()


def _read_summary_rows(ws_summary):
    total_row = None
    for row in range(1, ws_summary.max_row + 2):
        if str(ws_summary.cell(row=row, column=3).value or "").strip().lower() == "total":
            total_row = row
            break
    if total_row is None:
        total_row = ws_summary.max_row + 1

    rows = []
    columns = "ABCDEFGHIJKLMNOPQRSTUV"
    for r in range(5, total_row):
        if ws_summary.cell(row=r, column=2).value in (None, ""):
            continue
        row_data = {col: ws_summary.cell(row=r, column=idx).value for idx, col in enumerate(columns, start=1)}
        rows.append(row_data)
    return rows


def _money_short(value: float) -> str:
    if value >= 1_000_000:
        return f"${value / 1_000_000:g}M"
    if value >= 1_000:
        return f"${value / 1_000:g}k"
    return f"${value:g}"


COVERAGE_LABELS = [("J", "D"), ("K", "TPD"), ("L", "CI"), ("M", "Early CI"), ("O", "PA"), ("P", "Med Reimb")]


def _sum_assured_text(row: dict) -> str:
    groups = {}
    order = []
    for col, label in COVERAGE_LABELS:
        val = row.get(col)
        if isinstance(val, (int, float)) and val:
            if val not in groups:
                groups[val] = []
                order.append(val)
            groups[val].append(label)
    parts = [f"{_money_short(v)} {', '.join(groups[v])}" for v in order]
    di = row.get("N")
    if isinstance(di, (int, float)) and di:
        parts.append(f"{_money_short(di)}/mo DI")
    return "   ".join(parts) if parts else "Coverage not extracted"


def _premium_display(row: dict) -> str | None:
    cash = row.get("F") if isinstance(row.get("F"), (int, float)) else 0
    cpf = row.get("G") if isinstance(row.get("G"), (int, float)) else 0
    total = cash + cpf
    if not total:
        return None
    freq = str(row.get("H") or "").strip().lower()
    if freq == "monthly":
        return f"Premium ${total / 12:,.2f}/M"
    if freq == "quarterly":
        return f"Premium ${total / 4:,.2f}/Q"
    return f"Premium ${total:,.0f} p.a."


def _policy_label(row: dict) -> str:
    company = row.get("B") or ""
    plan_type = row.get("E") or ""
    policy_no = row.get("C") or ""
    return f"{company} {plan_type} ({policy_no})".strip()


def _policy_end_age(prot: dict, dob_date):
    """The single source of truth for a protection policy's real end age -
    used consistently everywhere an end age matters (axis milestones, bar
    length, whether a policy gets a fallback marker), so all three always
    agree on the same number for the same policy. Prefers the explicit
    coverage_end_age field (column T - set either from extraction, or by
    hand when the agent knows a fact the PDF didn't state); falls back to
    the maturity age already implied by the policy's own start-maturity
    date range (column D) if T wasn't captured. None if neither is known."""
    t = prot.get("T")
    if isinstance(t, (int, float)) and t > 0:
        return int(t)
    return _maturity_age_from_range_text(prot.get("D"), dob_date)


def _milestones(protection_rows: list[dict], dob_date=None) -> list[int]:
    """65 (CPF LIFE) is always a milestone. Beyond that, every protection
    policy's own real end age is used if we can determine one (see
    _policy_end_age). Ages differ client to client, so nothing here is
    hardcoded to any one client's numbers - a policy that ends at 70 puts
    70 on the axis, one that ends at 80 puts 80, and so on. Only falls back
    to generic spacing (75/85/95) when no policy gives us a real age."""
    ages = {65}
    derived = {
        age for r in protection_rows
        if (age := _policy_end_age(r, dob_date)) is not None and 65 < age <= 120
    }
    ages |= derived if derived else {75, 85, 95}
    return sorted(ages)


def _shade_range(ws, start_row: int, end_row: int, start_col: int, end_col: int, fill: PatternFill) -> None:
    if fill is NO_FILL or end_col < start_col:
        return
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).fill = fill


def _shade_block(ws, start_row: int, end_row: int, fill: PatternFill, width: int = BLOCK_WIDTH_COLS) -> None:
    _shade_range(ws, start_row, end_row, 1, width, fill)


# Fallback shading width for a Term plan when we can't compute an actual
# maturity age (e.g. no DOB on file for this client) - still visibly shorter
# than a permanent plan's full-width bar, signalling "this one lapses,"
# even without knowing exactly when. Deliberately NOT 0.6 - at
# BLOCK_WIDTH_COLS=10 that rounds to column 6, the same column age 65
# always occupies, which would make an unknown-end-age Term bar look like
# it deliberately ends at 65 (it doesn't - we just don't know its real end).
TERM_BAR_FRACTION = 0.5


def _dob_date_from_cell(value):
    """Same loose date parsing _current_age relies on, but returns the date
    itself (not just an age) so a policy's maturity age can be derived."""
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d-%b-%Y"):
        try:
            return datetime.strptime(str(value), fmt).date()
        except ValueError:
            continue
    return None


def _maturity_age_from_range_text(range_text, dob_date):
    """The Payment Date column (D) already stores a computed 'start - maturity'
    text. Parses the maturity half back into an age (given the client's DOB)
    so a policy's illustration bar can stop where its cover actually lapses,
    instead of always running the full width like a permanent plan."""
    if not isinstance(range_text, str) or dob_date is None or " - " not in range_text:
        return None
    maturity_date = _dob_date_from_cell(range_text.rsplit(" - ", 1)[-1].strip())
    if maturity_date is None:
        return None
    return maturity_date.year - dob_date.year


def _age_to_col(age, current_age) -> int:
    """Places any age on the axis by straight-line interpolation between
    just two fixed points: current age at column 3, and age 100 (the
    chart's practical right edge - see _bar_end_col) at BLOCK_WIDTH_COLS.
    This makes the scale genuinely proportional to real years - a 5-year
    gap is visually a fraction of a 30-year gap, not whatever a fixed,
    evenly-spaced label slot happened to allocate it. (An earlier version
    anchored every milestone to its own fixed column, 2 columns apart
    regardless of the actual age gap - which made e.g. a 65-to-70 span
    look as wide as a 70-to-100 span. This replaces that entirely.)"""
    if current_age is None:
        return BLOCK_WIDTH_COLS
    if age <= current_age:
        return 3
    if age >= 100:
        return BLOCK_WIDTH_COLS
    frac = (age - current_age) / (100 - current_age)
    col = 3 + frac * (BLOCK_WIDTH_COLS - 3)
    return max(3, min(BLOCK_WIDTH_COLS, round(col)))


def _milestone_columns(milestones, current_age):
    """Assigns each milestone age a column on the proportional scale from
    _age_to_col, then nudges any collision forward by one column - two
    milestones only round to the same column when they're genuinely close
    together relative to the whole current-age-to-100 span (e.g. 65 and 70
    for a young client), and silently dropping one of the two labels would
    be worse than a small, honest compromise on exact proportionality.
    Returns {age: column}, used both for the axis labels and for making
    sure any bar ending at one of these exact ages stops at the same
    column its label sits in."""
    if current_age is None:
        return {}
    result = {}
    last_col = 3
    for age in sorted(milestones):
        col = max(_age_to_col(age, current_age), last_col + 1)
        col = min(col, BLOCK_WIDTH_COLS)
        result[age] = col
        last_col = col
    return result


def _has_known_end_age(prot, dob_date, current_age) -> bool:
    """True only when this policy's end age is real data (see
    _policy_end_age) - not the generic Term fallback fraction, which is
    just a visual "this is shorter than permanent" placeholder with no
    actual age behind it."""
    return _policy_end_age(prot, dob_date) is not None and current_age is not None


def _bar_end_col(prot, dob_date, current_age, milestone_col_for=None):
    """How many of the block's BLOCK_WIDTH_COLS columns to shade, so the bar's
    length reflects how long this specific policy's cover actually runs -
    full width for a permanent plan, a shorter bar for a Term plan that lapses
    within the illustrated horizon. Whatever the source, the result is capped
    at age 100 (_age_to_col already clamps there) - illustrations don't
    meaningfully show coverage past that point, so a policy's real maturity
    data (or the full-width permanent-plan fallback, when we don't know its
    actual end age) is never allowed to shade further right than that. When
    this policy's real end age is also a printed milestone, the bar uses
    that milestone's own (possibly collision-nudged) column, so the bar
    never stops at a different column than the age number marking it."""
    maturity_age = _policy_end_age(prot, dob_date)

    if maturity_age is not None and current_age is not None:
        if milestone_col_for and maturity_age in milestone_col_for:
            return milestone_col_for[maturity_age]
        return _age_to_col(maturity_age, current_age)

    plan_type = str(prot.get("E") or "").strip().lower()
    if plan_type == "term":
        return max(3, round(BLOCK_WIDTH_COLS * TERM_BAR_FRACTION))
    return BLOCK_WIDTH_COLS


def _protection_bar_segments(prot, dob_date, current_age, milestone_col_for, full_fill, reduced_fill):
    """Returns [(start_col, end_col, fill), ...] describing how to shade a
    protection block. Normally a single full-length segment. If the policy's
    cover is documented as stepping down to a lower amount at a known age
    (coverage_drop_age / reduced_coverage_amount, columns U/V) AND the
    client's current age is known, the bar splits into a full-tone segment up
    to the drop and a lighter-tone segment from the drop to where cover ends
    - so a reduction in coverage is visible, not just a reduction in cover
    ending. Without a known current age there's no reliable axis to place the
    drop on, so it's skipped rather than guessed."""
    end_col = _bar_end_col(prot, dob_date, current_age, milestone_col_for)
    drop_age = prot.get("U")
    reduced_amount = prot.get("V")
    if not isinstance(drop_age, (int, float)) or reduced_amount in (None, "") or current_age is None:
        return [(1, end_col, full_fill)]

    drop_col = _age_to_col(drop_age, current_age)
    drop_col = max(1, min(end_col, drop_col))
    if drop_col <= 1:
        return [(1, end_col, reduced_fill)]
    if drop_col >= end_col:
        return [(1, end_col, full_fill)]
    return [(1, drop_col, full_fill), (drop_col, end_col, reduced_fill)]


def rebuild_illustration_sheet(client_name: str):
    """Regenerates the 'Policy Illustration <Client>' sheet in that client's
    workbook from the data already on their Policy Summary sheet. Safe to call
    every time a policy is added — it fully rebuilds the sheet from scratch."""
    path = _client_path(client_name)
    if not path.exists():
        raise PolicyWorkbookError(f"No workbook found yet for {client_name}")

    wb = openpyxl.load_workbook(path)
    if SHEET_NAME not in wb.sheetnames:
        raise PolicyWorkbookError(f"Workbook is missing a '{SHEET_NAME}' sheet")
    ws_summary = wb[SHEET_NAME]

    current_age = _current_age(ws_summary["K3"].value)
    dob_date = _dob_date_from_cell(ws_summary["K3"].value)
    rows = _read_summary_rows(ws_summary)
    protection_rows = [r for r in rows if _lane_for(r.get("E")) == "protection"]
    investment_rows = [r for r in rows if _lane_for(r.get("E")) == "investment"]
    milestones = _milestones(protection_rows, dob_date)
    milestone_col_for = _milestone_columns(milestones, current_age)

    sheet_name = _illustration_sheet_name(client_name)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    # Match Johnson's clean white background (gridlines off) rather than
    # Excel's default checkerboard.
    ws.sheet_view.showGridLines = False
    # This chart is wide (12 columns of timeline) and short - without an
    # explicit page setup it defaults to portrait/no-scaling, which splits
    # the row content mid-bar onto a second page in print preview, PDF
    # export, and some mobile viewers. That made the WOL bar's tail, the
    # age labels, and the whole CPF band look like they'd vanished, when
    # they were just pushed onto a page the reader wasn't looking at.
    # Landscape + fit-to-one-page-wide keeps the whole illustration together.
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 2
    # Every timeline column (C through L) is the SAME width. C and F used to
    # be widened to 30 to fit the policy-label and CPF description text -
    # but those columns also double as age-axis anchor points (current age
    # sits in C, age 65 in F), and giving them extra width silently stretches
    # whatever age gap lands on them - the 65-70 gap looked roughly 3x wider
    # than 70-100 even though it covers far fewer years. Text that needs
    # more room still overflows into the next empty cell exactly as the
    # policy label text (with its embedded line breaks) already does.
    for col in "CDEFGHIJKL":
        ws.column_dimensions[col].width = 12

    ws["A1"] = client_name
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    as_of = f"Financial Plan Illustration (All information as per {date.today().strftime('%B %Y')})"
    ws["C1"] = as_of
    ws["C1"].font = Font(bold=True, size=11)

    ws["B3"] = "SUM ASSURED"
    ws["B3"].alignment = Alignment(horizontal="right", vertical="center")
    ws["B3"].font = Font(size=10)

    # Stack blocks shortest-bar-first, so the longest-running policy sits in
    # the row right above the age axis - the bars read like a staircase
    # leading down into the axis line instead of a random jumble of lengths.
    protection_rows = sorted(
        protection_rows,
        key=lambda prot: _bar_end_col(prot, dob_date, current_age, milestone_col_for),
    )
    row = 4
    for i, prot in enumerate(protection_rows):
        block_start = row
        ws.cell(row=row, column=1, value=_sum_assured_text(prot))
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.merge_cells(start_row=block_start, start_column=1, end_row=block_start + 2, end_column=1)

        ws.cell(row=row, column=3, value=_policy_label(prot))
        premium_text = _premium_display(prot)
        if premium_text:
            ws.cell(row=row + 1, column=3, value=premium_text)
        remarks = prot.get("Q")
        if remarks:
            text = str(remarks)
            if len(text) > 140:
                text = text[:140].rsplit(" ", 1)[0] + "..."
            ws.cell(row=row + 2, column=3, value=text)
            # No wrap_text here - the timeline columns are now a uniform
            # width for the age axis to stay proportional (see the column
            # width comment above), which is too narrow to wrap this text
            # nicely. It overflows into the empty cells to the right
            # instead, the same way the policy label line above it does.
            ws.cell(row=row + 2, column=3).alignment = Alignment(vertical="top")

        full_fill, reduced_fill = (
            (BAND_FILL_PROTECTION, BAND_FILL_PROTECTION_REDUCED) if i % 2 == 0
            else (BAND_FILL_PROTECTION_ALT, BAND_FILL_PROTECTION_ALT_REDUCED)
        )
        segments = _protection_bar_segments(prot, dob_date, current_age, milestone_col_for, full_fill, reduced_fill)
        for seg_start, seg_end, fill in segments:
            _shade_range(ws, block_start, block_start + 2, seg_start, seg_end, fill)

        # Mark where this bar actually stops whenever we don't have a real
        # age to print for it (e.g. a Term plan with no stated end date) -
        # a small red arrow at the cut-off column, in the same column the
        # axis uses, so it's clear cover ends there even without an exact
        # age. A bar backed by real data doesn't need this - its end age is
        # already printed on the axis itself.
        end_col = _bar_end_col(prot, dob_date, current_age, milestone_col_for)
        if not _has_known_end_age(prot, dob_date, current_age) and end_col < BLOCK_WIDTH_COLS:
            marker_cell = ws.cell(row=block_start, column=end_col, value="▼")
            marker_cell.font = Font(bold=True, color="FFCC0000")
            marker_cell.alignment = Alignment(horizontal="center", vertical="center")

        row = block_start + 3

    # No blank spacer here - the axis sits directly under the last
    # protection block (closing the visible gap between them), the same
    # way the block rows themselves sit flush against each other.
    axis_row = row
    for c in range(1, BLOCK_WIDTH_COLS + 1):
        ws.cell(row=axis_row, column=c).border = Border(top=Side(style="thin"), bottom=Side(style="thin"))
    age_cell = ws.cell(row=axis_row, column=3, value=current_age if current_age is not None else "?")
    age_cell.alignment = Alignment(horizontal="left")
    # A milestone that lands exactly where a real (data-backed) policy bar
    # actually stops - e.g. age 100 marking the end of the longest-running
    # block - is right-aligned instead, so the number sits flush against
    # the edge of the coloured cell it's marking rather than floating at
    # its left side. Bars using the generic Term fallback are excluded
    # here since their end column isn't a real age match, just a fraction.
    bar_end_cols = {
        _bar_end_col(prot, dob_date, current_age, milestone_col_for)
        for prot in protection_rows
        if _has_known_end_age(prot, dob_date, current_age)
    }
    # milestone_col_for (computed earlier, before blocks were laid out - see
    # _milestone_columns) is the single source of truth for where each age
    # sits: proportional to real years, with any collision already nudged
    # apart so no label gets silently dropped. Every bar that ends at one
    # of these exact ages was already made to stop at this same column.
    for age, col in milestone_col_for.items():
        milestone_cell = ws.cell(row=axis_row, column=col, value=age)
        milestone_cell.alignment = Alignment(horizontal="right" if col in bar_end_cols else "left")
    row = axis_row + 1

    cpf_row = row
    ws.cell(row=cpf_row, column=1, value="CPF")
    ws.cell(row=cpf_row, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=cpf_row, start_column=1, end_row=cpf_row + 1, end_column=1)
    # CPF LIFE payouts only start at 65, so both the shading and the label
    # text begin at that exact column on the proportional age axis (not
    # column 1) - confirmed directly with the agent: the row should visually
    # start where CPF income actually starts, not run the full width like
    # the permanent-plan bars above it.
    age_65_col = milestone_col_for.get(65, _age_to_col(65, current_age) if current_age is not None else 6)
    # Wording matched to Johnson's own reference sheet. The dollar figure
    # is CPF Board's own published CPF LIFE Standard Plan estimate at the
    # Full Retirement Sum (see CPF_LIFE_FRS_MONTHLY_ESTIMATE above) - a
    # today's-dollars planning estimate, not a personalized projection for
    # this client's actual retirement age.
    ws.cell(row=cpf_row, column=age_65_col, value="CPF Life (assuming meeting Minimum Sum Scheme)")
    ws.cell(row=cpf_row, column=age_65_col).font = IMPORTANT_TEXT_FONT
    ws.cell(
        row=cpf_row + 1,
        column=age_65_col,
        value=f"Income: ${CPF_LIFE_FRS_MONTHLY_ESTIMATE:,.0f}* per month for life (today's CPF LIFE estimate)",
    )
    ws.cell(row=cpf_row + 1, column=age_65_col).font = IMPORTANT_TEXT_FONT
    _shade_range(ws, cpf_row, cpf_row + 1, age_65_col, BLOCK_WIDTH_COLS, BAND_FILL_INVESTMENT)
    row = cpf_row + 2 + 1  # +1 blank spacer

    for i, inv in enumerate(investment_rows):
        block_start = row
        ws.cell(row=block_start, column=1, value=_policy_label(inv))
        ws.cell(row=block_start, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        premium_text = _premium_display(inv)
        if premium_text:
            ws.cell(row=block_start, column=3, value=premium_text)

        at_65 = inv.get("S")
        if isinstance(at_65, (int, float)):
            ws.cell(row=block_start, column=6, value=f"Account Value {_money_short(at_65)} (@ 8% IRR, proj. age 65)")

        current_val = inv.get("R")
        if isinstance(current_val, (int, float)):
            ws.cell(row=block_start + 1, column=3, value=current_val)
            ws.cell(row=block_start + 1, column=3).number_format = '"$"#,##0'

        if i % 2 == 1:
            _shade_block(ws, block_start, block_start + 1, BAND_FILL_INVESTMENT)
        row = block_start + 2 + 1  # +1 blank spacer between blocks

    ws.cell(row=row, column=2, value="CASH VALUE")
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=2).font = Font(size=10)

    last_row = row
    for r in range(3, last_row + 1):
        ws.cell(row=r, column=2).border = Border(
            left=Side(style="thin", color="FF808080"),
            top=ws.cell(row=r, column=2).border.top,
            bottom=ws.cell(row=r, column=2).border.bottom,
        )

    wb.save(path)
    return path
