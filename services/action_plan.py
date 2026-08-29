"""Builds/refreshes a client's 'Action Plan' worksheet - one row per
coverage gap, phrased as a recommended next action rather than a flat
observation. Regenerated from the Policy Summary sheet's current totals
every time a policy is added, the same way policy_illustration.py
regenerates its own sheet from the same source data.
"""
from __future__ import annotations

from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from services.policy_workbook import (
    FIRST_DATA_ROW,
    PolicyWorkbookError,
    SHEET_NAME,
    _client_path,
    _find_total_row,
    _sync_to_onedrive,
    action_items_for,
)

ACTION_SHEET_NAME_PREFIX = "Action Plan "

HEADER_FILL = PatternFill(start_color="FF99CCFF", end_color="FF99CCFF", fill_type="solid")
THIN = Side(style="thin")
CELL_BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
TITLE_FONT = Font(name="Times New Roman", size=16, bold=True)
AS_OF_FONT = Font(name="Times New Roman", size=10, italic=True)
HEADER_FONT = Font(name="Times New Roman", size=12, bold=True)
BODY_FONT = Font(name="Times New Roman", size=11)


def _action_sheet_name(client_name: str) -> str:
    # Excel sheet names cap at 31 characters - truncate the same way the
    # illustration sheet does, so a long client name doesn't error out.
    return f"{ACTION_SHEET_NAME_PREFIX}{client_name}"[:31]


def rebuild_action_plan_sheet(client_name: str) -> list[dict]:
    """Regenerates the 'Action Plan <Client>' sheet from that client's
    current Policy Summary data and returns the action items ({"gap",
    "action"} dicts), so the caller can also send them straight to
    Telegram. Safe to call every time a policy is added - it fully
    rebuilds the sheet from scratch."""
    path = _client_path(client_name)
    if not path.exists():
        raise PolicyWorkbookError(f"No workbook found yet for {client_name}")

    wb = openpyxl.load_workbook(path)
    if SHEET_NAME not in wb.sheetnames:
        raise PolicyWorkbookError(f"Workbook is missing a '{SHEET_NAME}' sheet")
    ws_summary = wb[SHEET_NAME]
    total_row = _find_total_row(ws_summary)
    last_row = total_row - 1
    items = action_items_for(ws_summary, total_row, FIRST_DATA_ROW, last_row)

    sheet_name = _action_sheet_name(client_name)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 55

    ws["A1"] = f"Action Plan — {client_name}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")
    ws["A2"] = f"As of {date.today().strftime('%d %b %Y')}"
    ws["A2"].font = AS_OF_FONT
    ws.merge_cells("A2:B2")

    header_row = 4
    ws.cell(row=header_row, column=1, value="What's lacking")
    ws.cell(row=header_row, column=2, value="Recommended next action")
    for col in (1, 2):
        cell = ws.cell(row=header_row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row = header_row + 1
    for item in items:
        gap_cell = ws.cell(row=row, column=1, value=item["gap"])
        action_cell = ws.cell(row=row, column=2, value=item["action"])
        for cell in (gap_cell, action_cell):
            cell.font = BODY_FONT
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row += 1

    wb.save(path)
    _sync_to_onedrive(client_name, path)
    return items
